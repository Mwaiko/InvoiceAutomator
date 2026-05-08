"""
export_original_invoices.py  –  Export original (non-duplicate, non-credit-note)
KRA eTIMS invoices to Excel with full non-fiscal details.

What this does
──────────────
1.  Logs in and pages through the full receipt list (reuses fill_kra.py / search_duplicates.py).
2.  Keeps ONLY original invoices:
      • Skips credit notes  (amount < 0  OR  rcptType contains "Credit")
      • For duplicate groups (same amount + customer + date) keeps ONLY the first
        receipt; all later ones are duplicates and are dropped.
3.  For every kept invoice it opens the Invoice popup to collect:
      • Buyer PIN / TIN      (custTin)
      • Non-Fiscal Data      (remark) – e.g. "Order No, GRN No, Invoice No, Store No"
      • Total Taxable Amount (totTaxblAmt)
      • Total Tax Amount     (totTaxAmt)
      • Total Amount         (sumTotAmt)
4.  Writes a formatted .xlsx file.

Detail endpoint confirmed from browser Network capture:
  POST /app/ebm/trns/popup/popupTrnsSalesInvoice
  payload: tin=<KRA_PIN>&bhfId=<BRANCH>&invcNo=<receipt_no>
           &popId=popupInvoice&popTitle=Invoice+Information
           &popSize=XL&multiYn=N&callBackFnc=

Usage
─────
  # All time:
  python export_original_invoices.py --start 20210101 --end 20260507

  # Single date:
  python export_original_invoices.py --date 20260507

  # Today (default):
  python export_original_invoices.py

  # Skip detail fetches (faster – no PIN / non-fiscal data):
  python export_original_invoices.py --start 20260101 --end 20260507 --no-detail

  # Custom output path:
  python export_original_invoices.py --start 20260101 --end 20260507 --out invoices.xlsx

Env vars required (same as fill_kra.py):
  KRA_USERNAME, KRA_PASSWORD, KRA_PIN, KRA_BRANCH  (in .env or shell)
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import sys
import time
from datetime import date, datetime
from typing import NamedTuple

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

load_dotenv()

try:
    from fill_kra import EtimsConfig, KraError, _make_session, login
except ImportError:
    sys.exit(
        "❌  fill_kra.py not found in the same directory.\n"
        "    Place this script alongside fill_kra.py and try again."
    )

try:
    from search_duplicates import ReceiptRow, fetch_all_receipts, find_duplicates
except ImportError:
    sys.exit(
        "❌  search_duplicates.py not found in the same directory.\n"
        "    Place this script alongside search_duplicates.py and try again."
    )

log = logging.getLogger(__name__)

# Confirmed from browser Network capture
POPUP_PATH   = "/app/ebm/trns/popup/popupTrnsSalesInvoice"
LIST_REFERER = "/app/ebm/trns/sales/indexTrnsSalesReceipt"
TIMEOUT_S    = 30
DETAIL_DELAY = 0.3   # seconds between popup requests — be polite to the portal


# ─────────────────────────────────────────────────────────────────────────────
# FILTER: originals only
# ─────────────────────────────────────────────────────────────────────────────

def filter_originals(all_rows: list[ReceiptRow]) -> list[ReceiptRow]:
    """
    Return only 'original' invoices:
      • Drop credit notes (amount <= 0, or rcpt_type contains 'credit').
      • From each duplicate group keep only pos_rows[0] (the first / original).
    """
    positives = [
        r for r in all_rows
        if r.amount > 0 and "credit" not in r.rcpt_type.lower()
    ]

    duplicates = find_duplicates(all_rows)
    dup_nos: set[str] = set()
    for group_data in duplicates.values():
        for dup_row in group_data["positive"][1:]:
            dup_nos.add(dup_row.receipt_no)

    originals = [r for r in positives if r.receipt_no not in dup_nos]

    log.info(
        "Filtering: %d total -> %d positive -> %d after removing %d duplicates",
        len(all_rows), len(positives), len(originals), len(dup_nos),
    )
    return originals


# ─────────────────────────────────────────────────────────────────────────────
# DETAIL FETCH  (popup endpoint confirmed from browser)
# ─────────────────────────────────────────────────────────────────────────────

class InvoiceDetail(NamedTuple):
    receipt_no:     str
    cust_pin:       str    # Buyer KRA PIN / TIN  (custTin — may be blank)
    cust_name:      str    # Buyer name            (custNm from list view)
    rcpt_date:      str
    rcpt_type:      str
    taxable_amount: float  # totTaxblAmt
    vat_amount:     float  # totTaxAmt
    total_amount:   float  # sumTotAmt
    non_fiscal:     str    # remark field


def _clean(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def _input_val(soup: BeautifulSoup, name: str) -> str:
    tag = soup.find("input", {"name": name})
    return _clean(tag.get("value", "") if tag else "")


def _parse_amount(raw: str) -> float:
    try:
        return float(raw.replace(",", "").replace(" ", ""))
    except (ValueError, AttributeError):
        return 0.0


def fetch_detail(
    session:    requests.Session,
    cfg:        EtimsConfig,
    receipt_no: str,
) -> dict:
    """
    POST to popupTrnsSalesInvoice and return a dict of parsed fields.
    Returns {} on failure so export continues with list-view data only.
    """
    url  = f"{cfg.base_url}{POPUP_PATH}"
    hdrs = {
        "Accept":           "text/html, */*; q=0.01",
        "Content-Type":     "application/x-www-form-urlencoded; charset=UTF-8",
        "Origin":           cfg.base_url,
        "Referer":          f"{cfg.base_url}{LIST_REFERER}",
        "X-Requested-With": "XMLHttpRequest",
    }
    payload = {
        "tin":         cfg.pin,
        "bhfId":       cfg.branch,
        "invcNo":      receipt_no,
        "popId":       "popupInvoice",
        "popTitle":    "Invoice Information",
        "popSize":     "XL",
        "multiYn":     "N",
        "callBackFnc": "",
    }

    try:
        r = session.post(url, data=payload, headers=hdrs, timeout=TIMEOUT_S)
        if not r.ok:
            log.warning("Detail fetch failed for %s: HTTP %s", receipt_no, r.status_code)
            return {}
        return _parse_popup(r.text)
    except Exception as exc:
        log.warning("Detail fetch error for %s: %s", receipt_no, exc)
        return {}


def _parse_popup(html: str) -> dict:
    """
    Parse popupTrnsSalesInvoice HTML.

    Confirmed field names from the actual portal response:
      custTin     – Buyer PIN (may be blank for non-VAT buyers)
      remark      – Non-Fiscal Data (free-text reference string)
      totTaxblAmt – Total Taxable Amount
      totTaxAmt   – Total Tax Amount
      sumTotAmt   – Grand Total (also appears as sumTtAmt in some rows)
    """
    soup = BeautifulSoup(html, "html.parser")

    cust_pin   = _input_val(soup, "custTin")
    non_fiscal = _input_val(soup, "remark")

    # sumTtAmt is the readonly "original total" shown in the header summary table;
    # sumTotAmt is the running recalc total — both hold the same value here.
    taxable = _parse_amount(_input_val(soup, "totTaxblAmt"))
    vat     = _parse_amount(_input_val(soup, "totTaxAmt"))
    total   = _parse_amount(
        _input_val(soup, "sumTtAmt") or _input_val(soup, "sumTotAmt")
    )

    return {
        "cust_pin":       cust_pin,
        "non_fiscal":     non_fiscal,
        "taxable_amount": taxable,
        "vat_amount":     vat,
        "total_amount":   total,
    }


def enrich_rows(
    session:       requests.Session,
    cfg:           EtimsConfig,
    rows:          list[ReceiptRow],
    fetch_details: bool = True,
) -> list[InvoiceDetail]:
    results: list[InvoiceDetail] = []
    total = len(rows)

    for i, row in enumerate(rows, 1):
        log.info("  [%d/%d] receipt %s …", i, total, row.receipt_no)

        detail: dict = {}
        if fetch_details:
            detail = fetch_detail(session, cfg, row.receipt_no)
            time.sleep(DETAIL_DELAY)

        results.append(InvoiceDetail(
            receipt_no     = row.receipt_no,
            cust_pin       = detail.get("cust_pin",       ""),
            cust_name      = row.cust_name,
            rcpt_date      = row.rcpt_date,
            rcpt_type      = row.rcpt_type,
            taxable_amount = detail.get("taxable_amount", 0.0),
            vat_amount     = detail.get("vat_amount",     0.0),
            total_amount   = detail.get("total_amount",   0.0) or row.amount,
            non_fiscal     = detail.get("non_fiscal",     ""),
        ))

    return results


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

CLR_HEADER_BG = "1F4E79"
CLR_HEADER_FG = "FFFFFF"
CLR_ALT_ROW   = "EBF3FB"
CLR_TOTAL_BG  = "D6E4F0"
CLR_BORDER    = "AABDD6"

THIN = Side(style="thin", color=CLR_BORDER)
TBDR = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
AMT  = '#,##0.00'

COLUMNS = [
    ("Receipt No",          16),
    ("Buyer PIN",           18),
    ("Buyer Name",          34),
    ("Date",                14),
    ("Receipt Type",        16),
    ("Taxable Amt (KES)",   20),
    ("VAT (KES)",           16),
    ("Total Amt (KES)",     20),
    ("Non-Fiscal Data",     55),
]


def _hdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=True, color=CLR_HEADER_FG, size=10)
    c.fill      = PatternFill("solid", start_color=CLR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = TBDR


def _dat(ws, row, col, val, alt=False, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=10)
    c.border    = TBDR
    c.alignment = Alignment(vertical="center", wrap_text=(col == 9))
    if alt:
        c.fill = PatternFill("solid", start_color=CLR_ALT_ROW)
    if fmt:
        c.number_format = fmt


def export_to_excel(
    details:  list[InvoiceDetail],
    start_dt: str,
    end_dt:   str,
    out_path: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Original Invoices"
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    # Title
    last_col = chr(64 + len(COLUMNS))
    ws.merge_cells(f"A1:{last_col}1")
    period = start_dt if start_dt == end_dt else f"{start_dt}  ->  {end_dt}"
    tc = ws["A1"]
    tc.value     = f"KRA eTIMS — Original Invoices   |   Period: {period}"
    tc.font      = Font(name="Arial", bold=True, size=13, color=CLR_HEADER_BG)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column headers
    for ci, (name, width) in enumerate(COLUMNS, 1):
        _hdr(ws, 2, ci, name)
        ws.column_dimensions[chr(64 + ci)].width = width
    ws.row_dimensions[2].height = 36

    # Data rows
    DR = 3
    for i, inv in enumerate(details):
        r, alt = DR + i, (i % 2 == 1)
        _dat(ws, r, 1, inv.receipt_no,     alt)
        _dat(ws, r, 2, inv.cust_pin,       alt)
        _dat(ws, r, 3, inv.cust_name,      alt)
        _dat(ws, r, 4, inv.rcpt_date,      alt)
        _dat(ws, r, 5, inv.rcpt_type,      alt)
        _dat(ws, r, 6, inv.taxable_amount, alt, AMT)
        _dat(ws, r, 7, inv.vat_amount,     alt, AMT)
        _dat(ws, r, 8, inv.total_amount,   alt, AMT)
        _dat(ws, r, 9, inv.non_fiscal,     alt)
        ws.row_dimensions[r].height = 18

    # Totals row
    last  = DR + len(details) - 1
    tot_r = last + 1
    ws.merge_cells(f"A{tot_r}:E{tot_r}")
    lbl = ws.cell(row=tot_r, column=1, value="TOTALS")
    lbl.font      = Font(name="Arial", bold=True, size=10)
    lbl.fill      = PatternFill("solid", start_color=CLR_TOTAL_BG)
    lbl.alignment = Alignment(horizontal="right", vertical="center")
    lbl.border    = TBDR

    for col, letter in [(6, "F"), (7, "G"), (8, "H")]:
        c = ws.cell(row=tot_r, column=col, value=f"=SUM({letter}{DR}:{letter}{last})")
        c.font          = Font(name="Arial", bold=True, size=10)
        c.fill          = PatternFill("solid", start_color=CLR_TOTAL_BG)
        c.number_format = AMT
        c.border        = TBDR
        c.alignment     = Alignment(vertical="center")

    c9 = ws.cell(row=tot_r, column=9, value="")
    c9.fill   = PatternFill("solid", start_color=CLR_TOTAL_BG)
    c9.border = TBDR
    ws.row_dimensions[tot_r].height = 22

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 32

    summary = [
        ("Report generated",    datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Period",              period),
        ("Total invoices",      len(details)),
        ("Total taxable (KES)", f"=SUM('Original Invoices'!F{DR}:F{last})"),
        ("Total VAT (KES)",     f"=SUM('Original Invoices'!G{DR}:G{last})"),
        ("Total amount (KES)",  f"=SUM('Original Invoices'!H{DR}:H{last})"),
    ]
    for ri, (lbl_txt, val) in enumerate(summary, 2):
        lc = ws2.cell(row=ri, column=1, value=lbl_txt)
        lc.font   = Font(name="Arial", bold=True, size=10, color=CLR_HEADER_BG)
        lc.border = TBDR
        vc = ws2.cell(row=ri, column=2, value=val)
        vc.font   = Font(name="Arial", size=10)
        vc.border = TBDR
        if isinstance(val, str) and val.startswith("="):
            vc.number_format = AMT

    wb.save(out_path)
    log.info("Saved %d invoices -> %s", len(details), out_path)


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def _to_portal_date(s: str) -> str:
    try:
        return datetime.strptime(s, "%Y%m%d").strftime("%d/%m/%Y")
    except ValueError:
        sys.exit(f"Invalid date '{s}'. Use YYYYMMDD, e.g. 20260507")


def main() -> None:
    today = date.today().strftime("%Y%m%d")

    p = argparse.ArgumentParser(
        description="Export original KRA eTIMS invoices to Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--date",      default=today,
                   help="Single date (YYYYMMDD, default: today)")
    p.add_argument("--start",     default=None,
                   help="Start of date range (YYYYMMDD). Overrides --date.")
    p.add_argument("--end",       default=None,
                   help="End of date range (YYYYMMDD). Overrides --date.")
    p.add_argument("--out",       default=None,
                   help="Output .xlsx path (default: original_invoices_<date>.xlsx)")
    p.add_argument("--no-detail", action="store_true",
                   help="Skip detail popup fetches (faster, no PIN/non-fiscal data).")
    p.add_argument("--cust-pin",  default="",
                   help="Filter list by customer KRA PIN (optional).")
    p.add_argument("-v", "--verbose", action="store_true",
                   help="Enable debug logging.")
    args = p.parse_args()

    logging.basicConfig(
        level   = logging.DEBUG if args.verbose else logging.INFO,
        format  = "%(asctime)s  %(levelname)-7s  %(message)s",
        datefmt = "%H:%M:%S",
    )

    username = os.environ.get("KRA_USERNAME", "")
    password = os.environ.get("KRA_PASSWORD", "")
    if not username or not password:
        sys.exit("KRA_USERNAME and KRA_PASSWORD must be set in .env or shell.")

    cfg = EtimsConfig(
        pin      = os.environ.get("KRA_PIN",    ""),
        username = username,
        password = password,
        branch   = os.environ.get("KRA_BRANCH", "00"),
    )

    if args.start and args.end:
        start_p, end_p = _to_portal_date(args.start), _to_portal_date(args.end)
        lbl_s,   lbl_e = args.start, args.end
    else:
        start_p = end_p = _to_portal_date(args.date)
        lbl_s   = lbl_e = args.date

    out_path = args.out or f"original_invoices_{lbl_s}_{lbl_e}.xlsx"

    session = _make_session()
    log.info("Logging in ...")
    login(cfg, session)

    log.info("Fetching receipt list  %s -> %s ...", start_p, end_p)
    all_rows = fetch_all_receipts(
        session  = session,
        cfg      = cfg,
        start_dt = start_p,
        end_dt   = end_p,
        cust_pin = args.cust_pin,
    )

    if not all_rows:
        print("\nNo receipts found for this date range.")
        sys.exit(0)

    originals = filter_originals(all_rows)

    if not originals:
        print("\nNo original invoices found after filtering.")
        sys.exit(0)

    details = enrich_rows(
        session       = session,
        cfg           = cfg,
        rows          = originals,
        fetch_details = not args.no_detail,
    )

    export_to_excel(details=details, start_dt=lbl_s, end_dt=lbl_e, out_path=out_path)
    print(f"\nDone. {len(details)} original invoices -> {out_path}")


if __name__ == "__main__":
    main()